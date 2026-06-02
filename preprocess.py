"""
PRISM ISM Dataset 数据预处理脚本
输出：data/ 目录下的6个JSON文件

政策文件提取策略（按文件类型）：
  PDF  → pypdf 直接提取；若字符数<50则判定为扫描版，回退到 pytesseract OCR
  HTML → BeautifulSoup 静态提取；若字符数<200则判定为 JS 渲染，回退到 playwright
"""

import json
import re
import os
from pathlib import Path
from collections import defaultdict

import openpyxl
from pypdf import PdfReader
from bs4 import BeautifulSoup

# ── 可选依赖（缺失时静默降级）──────────────────────────────────────────────────
try:
    from pdf2image import convert_from_path
    import pytesseract
    pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    HAS_OCR = True
except ImportError:
    HAS_OCR = False

try:
    from playwright.sync_api import sync_playwright
    HAS_PLAYWRIGHT = True
except ImportError:
    HAS_PLAYWRIGHT = False

BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
POLICY_DIR = BASE_DIR / "policy_files" / "policy_files"
EXCEL_PATH = BASE_DIR / "2023.12 ISM Dataset(1).xlsx"

DATA_DIR.mkdir(exist_ok=True)

# ─────────────────────────────────────────────
# 工具函数
# ─────────────────────────────────────────────

def parse_timeframe(val):
    """'130 days' → 130, 'none' → None"""
    if not val or str(val).strip().lower() == "none":
        return None
    m = re.match(r"(\d+)", str(val))
    return int(m.group(1)) if m else None

def normalize_coverage(val):
    """统一 Coverage 字段的大小写和拼写"""
    if not val:
        return None
    v = str(val).strip()
    mapping = {
        "cross sectoral": "Cross-sectoral",
        "cross-sectoral": "Cross-sectoral",
        "mixed": "Mixed",
        "sectoral": "Sectoral",
        "asset": "Asset-based",
        "asset based": "Asset-based",
        "draft form": "Draft",
        "draft": "Draft",
        "passed but not implemented": "Passed (not implemented)",
    }
    return mapping.get(v.lower(), v)

def normalize_notification(val):
    if not val:
        return None
    v = str(val).strip().lower()
    if v in ("mandatory", "mandatory for all sectors"):
        return "Mandatory (all)"
    if "some" in v or "certain" in v:
        return "Mandatory (some)"
    if "real estate" in v:
        return "Mandatory (real estate only)"
    if v == "none":
        return "Not mandatory"
    return str(val).strip()

def normalize_preapproval(val):
    return normalize_notification(val)

def safe_int(val):
    try:
        return int(val)
    except (TypeError, ValueError):
        return None

def safe_float(val):
    try:
        f = float(val)
        return round(f, 4)
    except (TypeError, ValueError):
        return None

def safe_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


# ─────────────────────────────────────────────
# 1. Time Series Data
# ─────────────────────────────────────────────

def process_timeseries(wb):
    print("处理 Time Series Data ...")
    ws = wb["Time Series Data"]
    headers = [cell.value for cell in ws[1]]

    SECTOR_COLS = list(range(34, 71))   # 34-70: 37个细分行业
    AGGR_COLS   = list(range(72, 80))   # 72-79: 8个聚合分类

    sector_names = [headers[i] for i in SECTOR_COLS if headers[i]]
    aggr_names   = [headers[i] for i in AGGR_COLS   if headers[i]]

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        # 行业细分
        sectors = {}
        for i in SECTOR_COLS:
            h = headers[i]
            if h:
                sectors[h] = 1 if row[i] == 1 else 0

        # 行业聚合
        aggr = {}
        for i in AGGR_COLS:
            h = headers[i]
            if h:
                aggr[h] = safe_float(row[i])

        rec = {
            "country":              safe_str(row[0]),
            "iso3n":                safe_int(row[1]),
            "oecd":                 safe_int(row[2]),
            "eu_eea":               safe_int(row[3]),
            "five_eyes":            safe_int(row[4]),
            "year_established":     safe_int(row[5]),
            "year":                 safe_int(row[6]),
            "new_law_this_year":    1 if row[7] and str(row[7]).strip().lower() == "yes" else 0,
            "num_law_changes":      safe_int(row[8]),
            "num_mechanisms":       safe_int(row[9]),
            "coverage":             normalize_coverage(row[10]),
            "notification":         normalize_notification(row[11]),
            "preapproval":          normalize_preapproval(row[12]),
            "formal_call_in":       safe_int(row[13]),
            "greenfield_covered":   safe_int(row[14]),
            "real_estate_covered":  safe_int(row[15]),
            "time_frame_days":      parse_timeframe(row[16]),
            "threshold":            safe_float(row[17]),
            "review_increased_ownership": safe_int(row[18]),
            "filing_fees":          safe_int(row[19]),
            "mitigation":           safe_int(row[20]),
            "fines":                safe_int(row[21]),
            "ns_test":              safe_int(row[22]),
            "net_benefit_test":     safe_int(row[23]),
            "competition_test":     safe_int(row[24]),
            "interagency_review":   safe_int(row[25]),
            "lead_authority":       safe_str(row[26]),
            "tiered_authority":     safe_int(row[27]),
            "local_representation": safe_int(row[28]),
            "enhanced_gov_control": safe_int(row[29]),
            "colocation":           safe_int(row[30]),
            "export_control_dualuse": safe_int(row[31]),
            "eu_noeu_diff":         safe_float(row[32]),
            "covid_temporary":      safe_int(row[33]),
            "total_sectors":        safe_int(row[70]),
            "sectors":              sectors,
            "sector_categories":    aggr,
        }
        records.append(rec)

    out = {
        "meta": {
            "total_rows": len(records),
            "sector_names": sector_names,
            "sector_category_names": aggr_names,
        },
        "data": records
    }
    path = DATA_DIR / "timeseries.json"
    path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  → {path} ({len(records)} 条)")
    return records


# ─────────────────────────────────────────────
# 2. Event Data（法规级别）
# ─────────────────────────────────────────────

def process_events(wb):
    print("处理 Event Data ...")
    ws = wb["Event Data"]
    headers = [cell.value for cell in ws[1]]

    SECTOR_COLS = list(range(30, 67))

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[2]:
            continue

        sectors = {}
        for i in SECTOR_COLS:
            h = headers[i]
            if h:
                sectors[h] = 1 if row[i] == 1 else 0

        # Year Signed/Effective 可能是字符串或整数
        year_signed    = safe_int(row[3]) or (int(row[3]) if row[3] and str(row[3]).isdigit() else None)
        year_effective = safe_int(row[4]) or (int(row[4]) if row[4] and str(row[4]).isdigit() else None)

        rec = {
            "no":               safe_int(row[0]),
            "file_name":        safe_str(row[1]),
            "country":          safe_str(row[2]),
            "year_signed":      safe_int(str(row[3]).strip()) if row[3] else None,
            "year_effective":   safe_int(str(row[4]).strip()) if row[4] else None,
            "regulation":       safe_str(row[5]),
            "coverage":         normalize_coverage(row[6]),
            "notification":     normalize_notification(row[7]),
            "preapproval":      normalize_preapproval(row[8]),
            "formal_call_in":   safe_int(row[9]),
            "greenfield":       safe_int(row[10]),
            "real_estate":      safe_int(row[11]),
            "time_frame_days":  parse_timeframe(row[12]),
            "threshold":        safe_float(row[13]),
            "review_increased": safe_int(row[14]),
            "filing_fees":      safe_int(row[15]),
            "mitigation":       safe_int(row[16]),
            "fines":            safe_int(row[17]),
            "ns_test":          safe_int(row[18]),
            "net_benefit_test": safe_int(row[19]),
            "competition_test": safe_int(row[20]),
            "interagency":      safe_int(row[21]),
            "lead_authority":   safe_str(row[22]),
            "tiered_authority": safe_int(row[23]),
            "local_rep":        safe_int(row[24]),
            "enhanced_gov":     safe_int(row[25]),
            "colocation":       safe_int(row[26]),
            "export_dualuse":   safe_int(row[27]),
            "eu_noeu_diff":     safe_float(row[28]),
            "covid_temporary":  safe_int(row[29]),
            "total_sectors":    safe_int(row[66]) if len(row) > 66 else None,
            "sectors":          sectors,
        }
        records.append(rec)

    # 按国家分组，方便前端使用
    by_country = defaultdict(list)
    for r in records:
        if r["country"]:
            by_country[r["country"].strip()].append(r)

    out = {
        "meta": {"total_rows": len(records)},
        "data": records,
        "by_country": dict(by_country),
    }
    path = DATA_DIR / "events.json"
    path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  → {path} ({len(records)} 条)")
    return records


# ─────────────────────────────────────────────
# 3. List of all Changes
# ─────────────────────────────────────────────

def process_changes(wb):
    print("处理 List of all Changes ...")
    ws = wb["List of all Changes"]

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        country = safe_str(row[0])
        year    = safe_int(row[1])
        if not country or not year:
            continue

        rec = {
            "country":             country,
            "year":                year,
            "new_law":             1 if row[2] else 0,
            "new_amendment":       1 if row[3] else 0,
            "new_eo":              1 if row[4] else 0,
            "implementation_year": safe_int(row[5]),
            "superceded":          safe_str(row[6]),
            "explanation":         safe_str(row[7]),
            "name":                safe_str(row[8]),
        }
        records.append(rec)

    # 按国家分组
    by_country = defaultdict(list)
    for r in records:
        by_country[r["country"]].append(r)

    out = {
        "meta": {"total_rows": len(records)},
        "data": records,
        "by_country": dict(by_country),
    }
    path = DATA_DIR / "changes.json"
    path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  → {path} ({len(records)} 条)")
    return records


# ─────────────────────────────────────────────
# 4. yrly_new
# ─────────────────────────────────────────────

def process_yearly_new(wb):
    print("处理 yrly_new ...")
    ws = wb["yrly_new"]

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        country = safe_str(row[0])
        year    = safe_int(row[1])
        if not country or not year:
            continue
        rec = {
            "country":        country,
            "year":           year,
            "new_law":        safe_int(row[2]) or 0,
            "new_amendment":  safe_int(row[3]) or 0,
            "new_eo":         safe_int(row[4]) or 0,
            "new_reg_impl":   safe_int(row[5]) or 0,
            "total_new":      safe_int(row[6]) or 0,
        }
        records.append(rec)

    # 同时生成全球年度汇总（用于模块2时序图）
    global_yearly = defaultdict(lambda: {"new_law": 0, "new_amendment": 0, "new_eo": 0, "new_reg_impl": 0, "total_new": 0})
    for r in records:
        y = r["year"]
        global_yearly[y]["new_law"]       += r["new_law"]
        global_yearly[y]["new_amendment"] += r["new_amendment"]
        global_yearly[y]["new_eo"]        += r["new_eo"]
        global_yearly[y]["new_reg_impl"]  += r["new_reg_impl"]
        global_yearly[y]["total_new"]     += r["total_new"]

    out = {
        "meta": {"total_rows": len(records)},
        "data": records,
        "global_yearly": {str(k): v for k, v in sorted(global_yearly.items())},
    }
    path = DATA_DIR / "yearly_new.json"
    path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  → {path} ({len(records)} 条)")
    return records


# ─────────────────────────────────────────────
# 5. 政策文件文本提取
# ─────────────────────────────────────────────

def extract_pdf_text(filepath):
    """pypdf 直接提取；若字符数不足则 OCR 回退（需要 pytesseract + pdf2image）。"""
    try:
        reader = PdfReader(str(filepath), strict=False)
        pages = []
        for page in reader.pages[:30]:
            try:
                text = page.extract_text()
                if text and text.strip():
                    pages.append(text.strip())
            except Exception:
                continue
        text = "\n\n".join(pages)
    except Exception as e:
        return f"[PDF提取失败: {e}]"

    # 字符数不足 → 扫描版 PDF，尝试 OCR
    if len(text.strip()) < 50:
        text = _ocr_pdf(filepath)
    return text


POPPLER_PATH = r"C:\poppler\Library\bin"

def _ocr_pdf(filepath):
    """用 pytesseract 对扫描版 PDF 做 OCR（每页限前10页）。"""
    if not HAS_OCR:
        return "[OCR不可用: 请安装 pdf2image 和 pytesseract]"
    try:
        images = convert_from_path(str(filepath), last_page=10, dpi=200, poppler_path=POPPLER_PATH)
        pages = []
        for img in images:
            t = pytesseract.image_to_string(img, lang="eng")
            if t.strip():
                pages.append(t.strip())
        result = "\n\n".join(pages)
        return result if result.strip() else "[OCR结果为空]"
    except Exception as e:
        return f"[OCR失败: {e}]"


def extract_html_text(filepath):
    """BeautifulSoup 静态提取；若内容不足则 Playwright 渲染回退。"""
    try:
        content = Path(filepath).read_text(encoding="utf-8", errors="replace")
        soup = BeautifulSoup(content, "lxml")
        for tag in soup(["script", "style", "nav", "header", "footer"]):
            tag.decompose()
        main = soup.find("main") or soup.find("article") or soup.find("div", class_=re.compile(r"content|body|text", re.I))
        target = main if main else soup.body if soup.body else soup
        text = target.get_text(separator="\n")
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        text = "\n".join(lines)
    except Exception as e:
        return f"[HTML提取失败: {e}]"

    # 内容不足 → JS 动态渲染，尝试 Playwright
    if len(text.strip()) < 200:
        text = _playwright_html(filepath)
    return text


def _playwright_html(filepath):
    """用 Playwright 无头浏览器渲染 HTML 后提取正文。"""
    if not HAS_PLAYWRIGHT:
        return "[Playwright不可用: 请运行 pip install playwright && playwright install chromium]"
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()
            page.goto(filepath.as_uri(), wait_until="networkidle", timeout=30000)
            # 等待常见正文容器出现
            for sel in ["main", "article", "#content", ".content", "body"]:
                try:
                    page.wait_for_selector(sel, timeout=3000)
                    break
                except Exception:
                    continue
            text = page.inner_text("body")
            browser.close()
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        return "\n".join(lines)
    except Exception as e:
        return f"[Playwright失败: {e}]"

def chunk_text(text, chunk_size=500, overlap=100):
    """将长文本分成带重叠的小段，方便搜索和RAG"""
    words = text.split()
    chunks = []
    i = 0
    while i < len(words):
        chunk = " ".join(words[i:i + chunk_size])
        chunks.append(chunk)
        i += chunk_size - overlap
    return chunks

def parse_filename(filename):
    """'Australia_Foreign Acquisitions Act 1975.pdf' → {country, title}"""
    stem = Path(filename).stem
    parts = stem.split("_", 1)
    country = parts[0].strip() if len(parts) > 0 else "Unknown"
    title   = parts[1].strip() if len(parts) > 1 else stem
    return country, title

def process_policy_files():
    print("处理政策文件文本提取 ...")
    files = list(POLICY_DIR.glob("*.pdf")) + list(POLICY_DIR.glob("*.html"))
    print(f"  发现 {len(files)} 个文件")

    docs = []
    failed = []

    for f in sorted(files):
        country, title = parse_filename(f.name)
        ext = f.suffix.lower()

        if ext == ".pdf":
            text = extract_pdf_text(f)
        else:
            text = extract_html_text(f)

        if text.startswith("[") and "失败" in text:
            failed.append(f.name)
            print(f"  FAIL {f.name}")
            continue

        # 截断：保留前 8000 字符用于全文搜索
        text_truncated = text[:8000] if len(text) > 8000 else text
        chunks = chunk_text(text_truncated, chunk_size=300, overlap=50)

        # 标记提取方式，方便 debug
        if ext == ".pdf":
            method = "ocr" if "[OCR" not in text and len(text) > 50 and HAS_OCR else "pypdf"
        else:
            method = "playwright" if "[Playwright" not in text and len(text) > 200 and HAS_PLAYWRIGHT else "bs4"

        doc = {
            "id":       f.stem,
            "filename": f.name,
            "country":  country,
            "title":    title,
            "ext":      ext,
            "extract_method": method,
            "full_text_preview": text[:500],
            "chunks":   chunks,
            "char_count": len(text),
        }
        docs.append(doc)
        print(f"  OK [{method:10s}] {f.name[:60]} ({len(text)} chars, {len(chunks)} chunks)")

    # policy_texts.json：含完整 chunks，供 AI RAG 使用
    texts_path = DATA_DIR / "policy_texts.json"
    texts_path.write_text(
        json.dumps({"meta": {"total": len(docs), "failed": failed}, "docs": docs},
                   ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    # search_index.json：供 lunr.js 使用（不含 chunks，只含 id/title/country/preview）
    search_docs = [
        {
            "id":      d["id"],
            "filename": d["filename"],
            "country": d["country"],
            "title":   d["title"],
            "body":    " ".join(d["chunks"][:5]),  # 前5段作为搜索内容
        }
        for d in docs
    ]
    search_path = DATA_DIR / "search_index.json"
    search_path.write_text(
        json.dumps({"docs": search_docs}, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )

    print(f"  → {texts_path} ({len(docs)} 个文档)")
    print(f"  → {search_path} (lunr索引源)")
    if failed:
        print(f"  FAILED {len(failed)}: {failed}")
    return docs


# ─────────────────────────────────────────────
# 6. 生成国家元数据（汇总各数据源，供地图和筛选用）
# ─────────────────────────────────────────────

def build_country_meta(ts_records, event_records, changes_records):
    print("生成国家元数据 ...")
    country_map = {}

    for r in ts_records:
        c = r["country"]
        if not c:
            continue
        if c not in country_map:
            country_map[c] = {
                "country": c,
                "iso3n": r["iso3n"],
                "oecd": r["oecd"],
                "eu_eea": r["eu_eea"],
                "five_eyes": r["five_eyes"],
                "year_established": r["year_established"],
                "years_in_dataset": [],
            }
        country_map[c]["years_in_dataset"].append(r["year"])

    # 附上法规数量
    reg_count = defaultdict(int)
    for e in event_records:
        if e["country"]:
            reg_count[e["country"].strip()] += 1

    change_count = defaultdict(int)
    for ch in changes_records:
        change_count[ch["country"]] += 1

    for c, meta in country_map.items():
        meta["regulation_count"] = reg_count.get(c, 0)
        meta["change_count"] = change_count.get(c, 0)
        meta["year_min"] = min(meta["years_in_dataset"]) if meta["years_in_dataset"] else None
        meta["year_max"] = max(meta["years_in_dataset"]) if meta["years_in_dataset"] else None
        del meta["years_in_dataset"]

    countries = sorted(country_map.values(), key=lambda x: x["country"])
    path = DATA_DIR / "countries.json"
    path.write_text(json.dumps({"countries": countries}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  → {path} ({len(countries)} 个国家)")
    return countries


# ─────────────────────────────────────────────
# 主流程
# ─────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 60)
    print("PRISM ISM 数据预处理")
    print(f"  OCR support (scanned PDF): {'[OK] pytesseract+pdf2image' if HAS_OCR else '[NO] not installed'}")
    print(f"  JS render support (dynamic HTML): {'[OK] playwright' if HAS_PLAYWRIGHT else '[NO] not installed'}")
    print("=" * 60)

    wb = openpyxl.load_workbook(str(EXCEL_PATH))

    ts_records      = process_timeseries(wb)
    event_records   = process_events(wb)
    changes_records = process_changes(wb)
    yearly_records  = process_yearly_new(wb)

    policy_docs = process_policy_files()

    build_country_meta(ts_records, event_records, changes_records)

    print()
    print("=" * 60)
    print("完成！输出文件：")
    for f in sorted(DATA_DIR.iterdir()):
        size_kb = f.stat().st_size / 1024
        print(f"  {f.name:30s}  {size_kb:8.1f} KB")
    print("=" * 60)
